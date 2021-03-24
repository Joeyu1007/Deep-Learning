import cv2
import torch
from PIL import Image
import numpy as np
import scipy.misc
#import pandas as pd
import openpyxl
import math

image = Image.open('C:/Users/Adm/Desktop/corn/1080/good/_DSC2838_gray_result_good.jpg')
corn_img = np.array(image)
#print(corn_img)
new_corn_img_blur = cv2.medianBlur(corn_img,7)
#print(new_corn_img)
kernel = np.array([[-1,-1,-1],
                    [-1,9,-1],
                    [-1,-1,-1]])

good_corn_image = cv2.filter2D(new_corn_img_blur, -1, kernel)
# image = Image.fromarray(new_corn_img)
# image.show()
# print(new_corn_img)

ret,good_image = cv2.threshold(good_corn_image,5,255,0)

h = cv2.findContours(good_image,cv2.RETR_TREE,cv2.CHAIN_APPROX_NONE)
#提取轮廓
contours = h[0]

print(len(contours))
#数据清洗
coordinate_len_list = []
corn_border_list = []
for i in range(0,len(contours)):
    #print(len(contours[i]))
    if len(contours[i]) > 100:
        corn_border_list.append(contours[i])
        coordinate_len_list.append(len(contours[i]))
print(sorted(coordinate_len_list)[-28:-3])
#print(len(coordinate_len_list))
new_coordinate_len_list = sorted(coordinate_len_list)[-28:-3]


# #坐标点个数差值
# difference_list = []
# for i in range(len(coordinate_len_list)-1,1,-1):
#     difference_list.append(sorted(coordinate_len_list)[i] - sorted(coordinate_len_list)[i-1])
# print(difference_list)

# #找到最后一个点
# find_cut = []
# for i in range(0,len(difference_list)):
#     if difference_list[i] > 500:
#         find_cut.append(i)
# last_target = find_cut[-1:]
# print(last_target)
# #print(sorted(coordinate_len_list,reverse=True)[last_target[0]])
# #删除噪音
# new_coordinate_len_list = sorted(coordinate_len_list,reverse=True)[:last_target[0]+1]
# #print(new_coordinate_len_list)
# difference_list = []
# for i in range(0,len(new_coordinate_len_list)-1):
#     difference_list.append(new_coordinate_len_list[i] - new_coordinate_len_list[i+1])

# #正过来找到最后一个点
# find_cut = []
# for i in range(0,len(difference_list)):
#     if difference_list[i] > 400:
#         find_cut.append(i)
# #print(find_cut)
# last_target = find_cut[-1:]
# # print(last_target)
# if(last_target):
#     #删除连在一起的点
#     new_coordinate_len_list = sorted(new_coordinate_len_list,reverse=True)[last_target[0]+1:]
#     print(new_coordinate_len_list)

goodcorn_border_list = []
for i in range(0,len(corn_border_list)):
    if len(corn_border_list[i]) in sorted(new_coordinate_len_list):
        goodcorn_border_list.append(corn_border_list[i])

#画出轮廓
for i in range(0,len(goodcorn_border_list)):
    x, y, w, h = cv2.boundingRect(goodcorn_border_list[i]) 
    cv2.rectangle(good_corn_image, (x,y), (x+w,y+h), (255,0,0), 5)

#print(image)
# image = Image.fromarray(good_corn_image)
# image.show()
#image.save('C:/Users/Adm/Desktop/corn/6016/good/find_normal/_DSC2839_gray_result_good.jpg')


#print(good_corn_image[0][0])
#写入excel
def writewb():
    '''写入excel文件'''
    xlsx="C:/Users/Adm/Desktop/corn/1080/test.xlsx"
    wb=openpyxl.load_workbook(xlsx)
    sheet=wb.active   #获取当前的活动工作表
    print(sheet.max_row)
    row1=0
    row2=0
    row3=0
    print(sheet.max_column)

    ##第一种写入文件的方式：
    col=sheet.max_column
    #for k in range(0,len(good_corn)):
    for i in range(0,len(goodcorn_border_list)):
        for j in range(0,len(goodcorn_border_list[i])):
            #print(k)
            print(i + 1)
            #sheet.cell(row=row1+1,column=1,value=k)
            sheet.cell(row=row1+1,column=1,value=i+1)
            sheet.cell(row=row2+1,column=2,value=goodcorn_border_list[i][j][0][0])
            sheet.cell(row=row3+1,column=3,value=goodcorn_border_list[i][j][0][1])
            row1 = row1 + 1
            row2 = row2 + 1
            row3 = row3 + 1

    wb.save(xlsx)  #保存

writewb()