import cv2
import torch
from PIL import Image
import numpy as np
import scipy.misc
#import pandas as pd
import openpyxl
import math


# image = Image.open('C:/Users/Adm/Desktop/玉米/bad_res.jpg')
# corn_img = np.array(image)

# bad_corn_image = cv2.imread('C:/Users/Adm/Desktop/corn/6016/bad/_DSC2838_gray_result_bad.jpg',cv2.IMREAD_GRAYSCALE)
# #print(image)

bad_corn = []
for img_num in range(38,48):
    print("第"+ str(img_num) + "张图-------------------------------------------")
    bad_corn_image = cv2.imread('C:/Users/Adm/Desktop/corn/1080/bad/_DSC28' + str(img_num) + '_gray_result_bad.jpg',cv2.IMREAD_GRAYSCALE)
    #print(image)

    ret,bad_image = cv2.threshold(bad_corn_image,5,255,0)

    h = cv2.findContours(bad_image,cv2.RETR_TREE,cv2.CHAIN_APPROX_NONE)
    #提取轮廓
    contours = h[0]

    print("总点数：")
    print(len(contours))
    #数据清洗
    coordinate_len_list = []
    corn_border_list = []
    for i in range(0,len(contours)):
        #print(len(contours[i]))
        if len(contours[i]) > 100:
            corn_border_list.append(contours[i])
            coordinate_len_list.append(len(contours[i]))
    print("大于100像素点个数排序：")
    print(sorted(coordinate_len_list))
    #print(len(coordinate_len_list))
    new_coordinate_len_list = sorted(coordinate_len_list)[-28:-3]

    # #坐标点个数差值
    # difference_list = []
    # for i in range(len(coordinate_len_list)-1,1,-1):
    #     difference_list.append(sorted(coordinate_len_list)[i] - sorted(coordinate_len_list)[i-1])
    # print("大于100像素点个数之差：")
    # print(difference_list)

    # #找到最后一个点
    # find_cut = []
    # for i in range(0,len(difference_list)):
    #     if difference_list[i] > 400:
    #         find_cut.append(i)
    
    # new_coordinate_len_list = sorted(coordinate_len_list,reverse=True)
    # if(len(find_cut) > 1):
    #     last_target = find_cut[-1:]
    #     #print(last_target)
    #     #print(sorted(coordinate_len_list,reverse=True)[last_target[0]])
    #     #删除噪音
    #     new_coordinate_len_list = sorted(coordinate_len_list,reverse=True)[:last_target[0]+1]
    #     #print(new_coordinate_len_list)
 
    # difference_list = []
    # for i in range(0,len(new_coordinate_len_list)-1):
    #     difference_list.append(new_coordinate_len_list[i] - new_coordinate_len_list[i+1])

    # #正过来找到最后一个点
    # find_cut = []
    # for i in range(0,len(difference_list)):
    #     if difference_list[i] > 400:
    #         find_cut.append(i)
    # #print(find_cut)
    # last_target = find_cut[-1:]
    # # print(last_target)
    # if(last_target):
    #     #删除连在一起的点
    #     new_coordinate_len_list = sorted(new_coordinate_len_list,reverse=True)[last_target[0]+1:]
    #     print(new_coordinate_len_list)

    badcorn_border_list = []
    for i in range(0,len(corn_border_list)):
        if len(corn_border_list[i]) in sorted(new_coordinate_len_list):
            badcorn_border_list.append(corn_border_list[i])

    #画出轮廓
    for i in range(0,len(badcorn_border_list)):
        x, y, w, h = cv2.boundingRect(badcorn_border_list[i]) 
        cv2.rectangle(bad_corn_image, (x,y), (x+w,y+h), (255,0,0), 5)

    #print(image)
    image = Image.fromarray(bad_corn_image)
    #image.show()
    #image.save('C:/Users/Adm/Desktop/corn/6016/good/find_normal/_DSC28' + str(img_num) + '_gray_result_good.jpg')

    #writewb(goodcorn_border_list)
    bad_corn.append(badcorn_border_list)


#print(good_corn)
print(len(bad_corn))

#写入excel
def writewb():
    '''写入excel文件'''
    xlsx="C:/Users/Adm/Desktop/corn/1080/bad_coordinate.xlsx"
    wb=openpyxl.load_workbook(xlsx)
    sheet=wb.active   #获取当前的活动工作表
    print(sheet.max_row)
    row1=0
    row2=0
    row3=0
    row4=0
    print(sheet.max_column)

    ##第一种写入文件的方式：
    col=sheet.max_column
    for k in range(0,len(bad_corn)):
        for i in range(0,len(bad_corn[k])):
            for j in range(0,len(bad_corn[k][i])):
                #print(k)
                print(i + 1)
                sheet.cell(row=row1+1,column=1,value=k)
                sheet.cell(row=row2+1,column=2,value=i+1)
                sheet.cell(row=row3+1,column=3,value=bad_corn[k][i][j][0][0])
                sheet.cell(row=row4+1,column=4,value=bad_corn[k][i][j][0][1])
                row1 = row1 + 1
                row2 = row2 + 1
                row3 = row3 + 1
                row4 = row4 + 1

    wb.save(xlsx)  #保存

#writewb()

print('-------------------------------------------------------')
print(len(bad_corn))
#goodcorn_border_list = good_corn[1]
#print(len(goodcorn_border_list))
average_mse_list = []
for k in range(0,len(bad_corn)):
    for i in range(0,len(bad_corn[k])):
        x_list = []
        y_list = []
        for j in range(0,len(bad_corn[k][i])):
            x_list.append(bad_corn[k][i][j][0][0]/1080)
            y_list.append(bad_corn[k][i][j][0][1]/718)
            #print(badcorn_border_list[0][j][0][0])
            #print(badcorn_border_list[0][j][0][1])
        x_average = np.average(x_list)
        y_average = np.average(y_list)

        #计算当前图形的所有点到中心点距离
        distance_list = []
        for j in range(0,len(x_list)):
            distance = math.pow(x_list[j] - x_average, 2) + math.pow(y_list[j] - y_average, 2)
            distance_list.append(distance)
        #print(np.average(distance_list))
        #print(len(distance_list))

        #平均绝对偏差
        def MSE(y, t, n):
            return 1 / n * np.sum(abs(y - t))

        #平均距离
        
        #print(MSE(distance_list,np.average(distance_list),len(distance_list)))
        average_mse = MSE(distance_list,np.average(distance_list),len(distance_list))
        average_mse_list.append(average_mse)
    


print(average_mse_list)

'''写入excel文件'''
xlsx="C:/Users/Adm/Desktop/corn/1080/bad_mse.xlsx"
wb=openpyxl.load_workbook(xlsx)
sheet=wb.active   #获取当前的活动工作表
print(sheet.max_row)
row1=0
row2=0
print(sheet.max_column)

##第一种写入文件的方式：
col=sheet.max_column
for i in range(0,len(average_mse_list)):
    print(i + 1)
    sheet.cell(row=row1+1,column=1,value=i+1)
    sheet.cell(row=row2+1,column=2,value=average_mse_list[i])
    row1 = row1 + 1
    row2 = row2 + 1

wb.save(xlsx)  #保存

# ret,bad_image = cv2.threshold(bad_corn_image,5,255,0)
# #image = cv2.Canny(image,50,200)
# #print(image)
# # g = cv2.cvtColor(image,cv2.COLOR_BGR2GRAY)
# # print(g)

# h = cv2.findContours(bad_image,cv2.RETR_TREE,cv2.CHAIN_APPROX_NONE)
# #提取轮廓
# contours = h[0]

# print(len(contours))
# #数据清洗
# coordinate_len_list = []
# corn_border_list = []
# for i in range(0,len(contours)):
#     #print(len(contours[i]))
#     if len(contours[i]) > 700:
#         corn_border_list.append(contours[i])
#         coordinate_len_list.append(len(contours[i]))
# print(sorted(coordinate_len_list))
# print(len(coordinate_len_list))

# badcorn_border_list = []
# for i in range(0,len(corn_border_list)):
#     if len(corn_border_list[i]) in sorted(coordinate_len_list):
#         badcorn_border_list.append(corn_border_list[i])

# #画出轮廓
# for i in range(0,len(badcorn_border_list)):
#   x, y, w, h = cv2.boundingRect(badcorn_border_list[i]) 
#   cv2.rectangle(bad_corn_image, (x,y), (x+w,y+h), (255,0,0), 5)

# #print(image)
# image = Image.fromarray(bad_corn_image)
# image.show()

# #写入excel
# def writewb():
#     '''写入excel文件'''
#     xlsx="C:/Users/Adm/Desktop/corn/1.xlsx"
#     wb=openpyxl.load_workbook(xlsx)
#     sheet=wb.active   #获取当前的活动工作表
#     print(sheet.max_row)
#     row1=0
#     row2=0
#     row3=0
#     print(sheet.max_column)

#     ##第一种写入文件的方式：
#     col=sheet.max_column
#     for i in range(0,len(badcorn_border_list)):
#       for j in range(0,len(badcorn_border_list[i])):
#         print(i + 1)
#         sheet.cell(row=row1+1,column=1,value=i+1)
#         sheet.cell(row=row2+1,column=2,value=badcorn_border_list[i][j][0][0])
#         sheet.cell(row=row3+1,column=3,value=badcorn_border_list[i][j][0][1])
#         row1 = row1 + 1
#         row2 = row2 + 1
#         row3 = row3 + 1

#     wb.save(xlsx)  #保存

#writewb()

# average_mse_list = []
# for i in range(0,len(badcorn_border_list)):
#     x_list = []
#     y_list = []
#     for j in range(0,len(badcorn_border_list[i])):
#         x_list.append(badcorn_border_list[i][j][0][0])
#         y_list.append(badcorn_border_list[i][j][0][1])
#         #print(badcorn_border_list[0][j][0][0])
#         #print(badcorn_border_list[0][j][0][1])
#     x_average = np.average(x_list)
#     y_average = np.average(y_list)

#     distance_list = []
#     for j in range(0,len(x_list)):
#         distance = math.pow(x_list[j] - x_average, 2) + math.pow(y_list[j] - y_average, 2)
#         distance_list.append(distance)
#     #print(np.average(distance_list))
#     #print(len(distance_list))

#     #平均绝对偏差
#     def MSE(y, t, n):
#         return 1 / n * np.sum(abs(y - t))


#     print(MSE(distance_list,np.average(distance_list),len(distance_list)))
#     average_mse = MSE(distance_list,np.average(distance_list),len(distance_list))
#     average_mse_list.append(average_mse)

# print(average_mse_list)

# '''写入excel文件'''
# xlsx="C:/Users/Adm/Desktop/corn/2.xlsx"
# wb=openpyxl.load_workbook(xlsx)
# sheet=wb.active   #获取当前的活动工作表
# print(sheet.max_row)
# row1=0
# row2=0
# print(sheet.max_column)

# ##第一种写入文件的方式：
# col=sheet.max_column
# for i in range(0,len(average_mse_list)):
#     print(i + 1)
#     sheet.cell(row=row1+1,column=1,value=i+1)
#     sheet.cell(row=row2+1,column=2,value=average_mse_list[i])
#     row1 = row1 + 1
#     row2 = row2 + 1

# wb.save(xlsx)  #保存

